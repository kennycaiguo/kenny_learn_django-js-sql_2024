from django.shortcuts import render, redirect, HttpResponse
from day16app import models  # 导入models因为以后可能里面会有很多类不能一个一个导入
from day16app.utils.pagination import Pagination
from openpyxl import load_workbook


def dep_list(req):
    """显示部门列表"""
    dep_list = models.Department.objects.all()
    # 分页功能实现
    # 1.实例化分页组件
    page_obj = Pagination(req, dep_list, page_size=8)
    # 2.获取分页后的数据
    page_query_set = page_obj.page_queryset
    # 3.获取分页器的html字符串
    page_str = page_obj.gen_html()
    # 渲染数据
    return render(req, "dep_list.html", {"deps": page_query_set, "page_str": page_str})


def dep_add(req):  # 这个功能太简单了,只有一个字段,不需要ModelForm
    """新增部门"""
    # get
    if req.method == "GET":
        return render(req, "dep_add.html")
    # post
    title = req.POST.get("title")
    models.Department.objects.create(title=title)
    return redirect("/dep/list")


def dep_del(req):
    """删除部门"""
    nid = req.GET.get("nid")
    models.Department.objects.filter(id=nid).delete()
    return redirect("/dep/list")


def dep_edit(req, nid):
    # 处理get
    # 保证nid存在
    if not models.Department.objects.filter(id=nid).exists():
        return redirect("/dep/list")  # 方式1 重定向
        # return render(req,"error.html")   # 方式2 展示错误页面
    if req.method == "GET":
        dep = models.Department.objects.filter(id=nid).first()
        return render(req, "dep_edit.html", {"dep": dep})
    # 处理post
    # id T= req.POS.get("id")  # 在模板里面没有使用hidden字段,这里就不需要这一句
    title = req.POST.get("title")
    models.Department.objects.filter(id=nid).update(title=title)
    return redirect("/dep/list")


def dep_multiadd(req):
    # 1.获取用户上传的文件,得到一个文件对象
    exc = req.FILES.get("exc")
    print(exec)
    # 2.读取这个文件对象的数据并且保存
    wb = load_workbook(exc)  # 这个函数不仅仅可以传递一个文件名,也可以传递一个文件对象作为参数,就是这里的exc文件对象.
    # print(wb)
    sheet = wb.worksheets[0]
    # 3.我们这个文件笔记特殊,只有一列,我们只需要获取这一列的数据,然后把这一列的每一行的数据保存到数据库即可.注意必须从2开始因为第一行是表头,我们不需要
    # 这个包裹只有一列,所以比较简单.
    for row in sheet.iter_rows(min_row=2):
        title1 = row[0].value
        # 先查询一下有没有这个部门,如果有,就不创建
        if not models.Department.objects.filter(title=title1).exists():
            models.Department.objects.create(title=title1)
    return redirect("/dep/list/")
